# -*- coding: utf-8 -*-
"""Leitura das planilhas.

SOMENTE LEITURA. Este modulo abre todo workbook com read_only=True e nunca
importa nada capaz de salvar. Nao existe caminho de escrita na Fase 2.

O carregamento e feito uma vez e mantido em memoria; a releitura acontece
quando o arquivo muda no disco (mtime) ou sob pedido explicito.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import os
import threading
from typing import Any, Iterator, Optional

import openpyxl

from . import normalize as nz
from . import schema as sch
from .models import ESTADO_SEM_BAIXA as models_ESTADO_SEM_BAIXA
from .models import (
    Aviso,
    Indice,
    Lancamento,
    LoteReembolso,
    Nota,
    NotaDebito,
    ReembolsoManual,
)


class ErroDeLeitura(Exception):
    """Falha ao ler um arquivo obrigatorio."""


# ---------------------------------------------------------------------------
# Acesso aos arquivos
# ---------------------------------------------------------------------------


def _caminho(base: str, arquivo: str) -> str:
    return os.path.join(base, arquivo)


def _abrir(base: str, arquivo: str):
    """Abre um workbook em modo somente leitura, com valores calculados."""
    caminho = _caminho(base, arquivo)
    if not os.path.exists(caminho):
        raise ErroDeLeitura(f"Arquivo nao encontrado: {arquivo}")
    return openpyxl.load_workbook(caminho, data_only=True, read_only=True)


def _aba(wb, nome: str):
    """Resolve o nome da aba tolerando espaco no fim e diferenca de acento.

    Necessario porque o arquivo da Rafaela tem abas gravadas como
    'julho 2026 ' (com espaco), armadilha registrada na Fase 1.
    """
    if nome in wb.sheetnames:
        return wb[nome]
    alvo = nz.chave(nome)
    for real in wb.sheetnames:
        if nz.chave(real) == alvo:
            return wb[real]
    return None


def _linhas(ws, primeira: int = 1) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """Itera linhas de uma aba read_only, devolvendo (numero_da_linha, valores)."""
    for i, linha in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < primeira:
            continue
        yield i, linha


def _col(linha: tuple[Any, ...], indice_1based: int) -> Any:
    """Le uma coluna da tupla da linha, tolerando linha curta."""
    i = indice_1based - 1
    if i < 0 or i >= len(linha):
        return None
    return linha[i]


def _vazia(linha: tuple[Any, ...], ate: int = 12) -> bool:
    return all(_col(linha, c) is None for c in range(1, ate + 1))


# ---------------------------------------------------------------------------
# Faturamento
# ---------------------------------------------------------------------------


def _ler_faturamento(base: str, entidade: sch.Entidade, indice: Indice) -> None:
    wb = _abrir(base, entidade.arquivo)
    try:
        for nome_aba in wb.sheetnames:
            if nome_aba.strip() in sch.FAT_ABAS_NAO_MENSAIS:
                continue
            periodo = nz.mes_de_aba(nome_aba)
            if periodo is None:
                continue
            mes, ano = periodo
            ws = wb[nome_aba]
            c = sch.COLS_FAT

            # O bloco principal comeca na linha 2 e termina na primeira linha
            # totalmente vazia. O que vier depois e bloco secundario, com
            # layout diferente - registrado, mas nunca somado aos totais.
            fim_do_bloco_principal = False
            for numero_linha, linha in _linhas(ws, c.primeira_linha):
                if _vazia(linha):
                    fim_do_bloco_principal = True
                    continue

                cliente = nz.texto(_col(linha, c.cliente))
                bruto = nz.numero(_col(linha, c.valor_bruto))
                liquido = nz.numero(_col(linha, c.valor_liquido))

                if fim_do_bloco_principal:
                    # Layout deslocado: cliente em A, valores em B e C.
                    indice.notas.append(
                        Nota(
                            entidade=entidade.codigo,
                            aba=nome_aba,
                            linha=numero_linha,
                            numero=None,
                            cliente=nz.texto(_col(linha, 1)),
                            valor_bruto=nz.numero(_col(linha, 2)),
                            valor_liquido=nz.numero(_col(linha, 3)),
                            referencia=nz.texto(_col(linha, 4)),
                            observacoes=nz.texto(_col(linha, 5)),
                            data_recebimento=nz.data(_col(linha, 6)),
                            marcador_pendente=nz.contem_pendente(_col(linha, 6)),
                            nota_livre=None,
                            mes=mes,
                            ano=ano,
                            bloco_secundario=True,
                        )
                    )
                    continue

                if cliente is None and bruto is None and liquido is None:
                    continue

                celula_g = _col(linha, c.data_recebimento)
                indice.notas.append(
                    Nota(
                        entidade=entidade.codigo,
                        aba=nome_aba,
                        linha=numero_linha,
                        numero=nz.texto(_col(linha, c.nf)),
                        cliente=cliente,
                        valor_bruto=bruto,
                        valor_liquido=liquido,
                        referencia=nz.texto(_col(linha, c.referencia)),
                        observacoes=nz.texto(_col(linha, c.observacoes)),
                        data_recebimento=nz.data(celula_g),
                        marcador_pendente=nz.contem_pendente(celula_g),
                        nota_livre=nz.texto(_col(linha, c.nota_livre)),
                        mes=mes,
                        ano=ano,
                    )
                )

        _ler_mapa_cnpj(wb, entidade, indice)
    finally:
        wb.close()

    _avisar_blocos_secundarios(entidade, indice)


def _ler_mapa_cnpj(wb, entidade: sch.Entidade, indice: Indice) -> None:
    """Le a aba NOTAS: qual cliente vai para qual CNPJ."""
    ws = _aba(wb, sch.NOTAS_CNPJ.aba)
    if ws is None:
        return
    n = sch.NOTAS_CNPJ
    for _, linha in _linhas(ws, n.primeira_linha):
        for coluna, destino in (
            (n.col_novo_cnpj, "rafaela"),
            (n.col_intermediario, "intermediario"),
            (n.col_manter, "principal"),
        ):
            nome = nz.texto(_col(linha, coluna))
            if nome:
                indice.mapa_cnpj.setdefault(nz.radical_cliente(nome), destino)


def _avisar_blocos_secundarios(entidade: sch.Entidade, indice: Indice) -> None:
    """Avisa sobre os blocos fora do layout e sobre repeticao entre meses."""
    secundarias = [
        n
        for n in indice.notas
        if n.entidade == entidade.codigo and n.bloco_secundario
    ]
    if not secundarias:
        return

    por_aba: dict[str, list[Nota]] = {}
    for n in secundarias:
        por_aba.setdefault(n.aba, []).append(n)

    for aba, linhas in por_aba.items():
        total = sum(n.valor for n in linhas)
        indice.avisos.append(
            Aviso(
                severidade="atencao",
                arquivo=entidade.arquivo,
                aba=aba,
                mensagem=(
                    f"{len(linhas)} linhas fora do bloco principal (a partir "
                    f"da linha {min(n.linha for n in linhas)}), somando "
                    f"{nz.moeda(total)}. O layout não segue o cabeçalho da "
                    f"aba, então elas não entram nos totais."
                ),
            )
        )

    # Mesma cobranca repetida em meses diferentes.
    assinaturas: dict[tuple, list[str]] = {}
    for n in secundarias:
        chave_nota = (nz.radical_cliente(n.cliente), round(n.valor, 2), n.referencia)
        assinaturas.setdefault(chave_nota, []).append(n.aba)
    repetidas = {k: v for k, v in assinaturas.items() if len(set(v)) > 1}
    if repetidas:
        valor_total = sum(k[1] for k in repetidas)
        abas = sorted({a for v in repetidas.values() for a in v})
        indice.avisos.append(
            Aviso(
                severidade="critico",
                arquivo=entidade.arquivo,
                aba=", ".join(abas),
                mensagem=(
                    f"{len(repetidas)} cobranças idênticas aparecem em mais de "
                    f"um mês ({', '.join(abas)}), somando "
                    f"{nz.moeda(valor_total)}. Pode ser a mesma dívida "
                    f"arrastada de um mês para o outro. Não somei. "
                    f"Pergunta 7 do mapeamento, em aberto."
                ),
            )
        )


# ---------------------------------------------------------------------------
# Razoes bancarios
# ---------------------------------------------------------------------------


def _ler_razoes(base: str, indice: Indice) -> None:
    wb = _abrir(base, sch.ARQ_FLUXO)
    try:
        # Uma leitura por ABA, nao por conta: Santander e Omie Cash dividem a
        # mesma aba, e ler duas vezes duplicaria todos os lancamentos dela.
        vistas: set[str] = set()
        for conta in sch.CONTAS.values():
            if conta.aba in vistas:
                continue
            vistas.add(conta.aba)
            ws = _aba(wb, conta.aba)
            if ws is None:
                indice.avisos.append(
                    Aviso("critico", sch.ARQ_FLUXO, conta.aba, "Aba nao encontrada.")
                )
                continue
            totais = 0
            for numero_linha, linha in _linhas(ws, 2):
                if _vazia(linha):
                    continue
                descricao = nz.texto(_col(linha, conta.descricao))
                saida = nz.zero(_col(linha, conta.saida))
                entrada = nz.zero(_col(linha, conta.entrada))
                if saida == 0 and entrada == 0 and descricao is None:
                    continue

                remetente = nz.texto(_col(linha, conta.remetente))
                contrato = nz.texto(_col(linha, conta.contrato))
                data_lanc = nz.data(_col(linha, conta.data))
                banco = nz.texto(_col(linha, conta.banco))

                # Todo razao termina com uma linha de SUBTOTAL, e a aba
                # SANTANDER tem ainda um subtotal no meio. Elas so tem valor,
                # sem remetente, contrato, descricao, data nem banco. Se
                # entrassem como lancamento, dobrariam os totais.
                if not any((remetente, contrato, descricao, data_lanc, banco)):
                    totais += 1
                    continue

                rotulo = (
                    nz.texto(_col(linha, conta.reembolso))
                    if conta.reembolso
                    else None
                )
                efetiva = sch.conta_efetiva(conta.aba, banco, conta.codigo)
                indice.lancamentos.append(
                    Lancamento(
                        conta=conta.codigo,
                        aba=conta.aba,
                        linha=numero_linha,
                        remetente=remetente,
                        contrato=contrato,
                        descricao=descricao,
                        saida=saida,
                        entrada=entrada,
                        data=data_lanc,
                        banco=banco,
                        historico=nz.texto(_col(linha, conta.historico)),
                        rotulo_reembolso=rotulo,
                        natureza=sch.classificar_descricao(descricao),
                        conta_efetiva=efetiva,
                        espelho=sch.e_espelho(conta.aba, efetiva),
                    )
                )

            if totais:
                indice.avisos.append(
                    Aviso(
                        "info",
                        sch.ARQ_FLUXO,
                        conta.aba,
                        f"{totais} linha(s) de subtotal ignorada(s). Sao "
                        f"fórmulas de fechamento da aba, não lançamentos.",
                    )
                )

        _ler_notas_debito(wb, indice)
    finally:
        wb.close()

    _avisar_espelhos(indice)


def _avisar_espelhos(indice: Indice) -> None:
    """Avisa quando uma aba repete lancamentos que ja moram em outra.

    Nao afirma que sao copias: confere uma a uma contra a aba de origem, por
    data, valor e contraparte, e informa quantas de fato batem.
    """
    espelhos = [l for l in indice.lancamentos if l.espelho]
    if not espelhos:
        return

    def assinatura(l) -> tuple:
        return (l.data, round(l.saida, 2), round(l.entrada, 2), nz.chave(l.remetente))

    por_par: dict[tuple[str, str], list] = {}
    for l in espelhos:
        por_par.setdefault((l.aba, l.conta_efetiva), []).append(l)

    for (aba, codigo), linhas in por_par.items():
        conta = sch.CONTAS.get(codigo)
        rotulo = conta.rotulo if conta else codigo
        casa = conta.aba if conta else "?"

        originais = {
            assinatura(x)
            for x in indice.lancamentos
            if not x.espelho and x.aba == casa
        }
        confirmadas = [l for l in linhas if assinatura(l) in originais]
        soltas = [l for l in linhas if assinatura(l) not in originais]

        texto = (
            f"{len(linhas)} linhas do {rotulo} aparecem na aba {aba}, somando "
            f"{nz.moeda(sum(x.saida for x in linhas))} em saídas, mas a conta "
            f"tem aba própria ({casa}). Conferi uma a uma: "
            f"{len(confirmadas)} repetem lá a mesma data, valor e contraparte. "
            f"Tratei todas como cópia e deixei fora dos totais, para não "
            f"contar duas vezes."
        )
        if soltas:
            texto += (
                f" Atenção: {len(soltas)} não foram encontradas na aba de "
                f"origem ({nz.moeda(sum(x.saida for x in soltas))}). Se forem "
                f"movimento real, estão faltando lá."
            )
        indice.avisos.append(
            Aviso(
                severidade="critico" if soltas else "atencao",
                arquivo=sch.ARQ_FLUXO,
                aba=aba,
                mensagem=texto,
            )
        )


def _ler_notas_debito(wb, indice: Indice) -> None:
    ws = _aba(wb, sch.ND.aba)
    if ws is None:
        indice.avisos.append(
            Aviso("critico", sch.ARQ_FLUXO, sch.ND.aba, "Aba nao encontrada.")
        )
        return
    d = sch.ND
    for numero_linha, linha in _linhas(ws, d.primeira_linha):
        numero = nz.texto(_col(linha, d.nd))
        if numero is None:
            continue
        celula_pagamento = _col(linha, d.data_pagamento)
        data_pagamento = nz.data(celula_pagamento)
        indice.notas_debito.append(
            NotaDebito(
                aba=sch.ND.aba,
                linha=numero_linha,
                numero=numero,
                responsavel=nz.texto(_col(linha, d.responsavel)),
                cliente=nz.texto(_col(linha, d.cliente)),
                data_envio=nz.data(_col(linha, d.data_envio)),
                valor=nz.numero(_col(linha, d.valor)),
                data_pagamento=data_pagamento,
                texto_pagamento=(
                    nz.texto(celula_pagamento) if data_pagamento is None else None
                ),
                despesas=nz.texto(_col(linha, d.despesas)),
            )
        )

    sem_valor = [n for n in indice.notas_debito if n.valor is None]
    if sem_valor:
        indice.avisos.append(
            Aviso(
                "atencao",
                sch.ARQ_FLUXO,
                sch.ND.aba,
                f"{len(sem_valor)} nota(s) de débito sem valor preenchido "
                f"({', '.join(n.numero or '?' for n in sem_valor)}). "
                f"Ficam fora do total em aberto.",
            )
        )


# ---------------------------------------------------------------------------
# Reembolsos de guias
# ---------------------------------------------------------------------------


def _ler_reembolsos(base: str, indice: Indice) -> None:
    wb = _abrir(base, sch.ARQ_REEMBOLSOS)
    try:
        for ano, cfg in sch.GUIAS.items():
            ws = _aba(wb, cfg.aba)
            if ws is None:
                indice.avisos.append(
                    Aviso("critico", sch.ARQ_REEMBOLSOS, cfg.aba, "Aba nao encontrada.")
                )
                continue

            for numero_linha, linha in _linhas(ws, cfg.lote_primeira_linha):
                if numero_linha >= cfg.lote_ultima_linha:
                    break
                status = nz.texto(_col(linha, cfg.lote_status))
                valor = nz.numero(_col(linha, cfg.lote_valor))
                descricao = nz.texto(_col(linha, cfg.lote_descricao))
                if valor is None or status is None:
                    continue
                indice.lotes.append(
                    LoteReembolso(
                        ano=ano,
                        aba=cfg.aba,
                        linha=numero_linha,
                        status=status,
                        valor=valor,
                        data_recebimento=nz.data(_col(linha, cfg.lote_data)),
                        descricao=descricao,
                        numero_lote=nz.extrair_lote(descricao),
                        pendente=sch.status_e_pendente(status),
                        quitado=sch.status_e_quitado(status),
                    )
                )

            if cfg.manual_primeira_linha is None:
                continue
            for numero_linha, linha in _linhas(ws, cfg.manual_primeira_linha):
                status = nz.texto(_col(linha, cfg.manual_status))
                valor = nz.numero(_col(linha, cfg.manual_valor))
                if valor is None or status is None:
                    continue
                indice.manuais.append(
                    ReembolsoManual(
                        aba=cfg.aba,
                        linha=numero_linha,
                        status=status,
                        valor=valor,
                        data_recebimento=nz.data(_col(linha, cfg.manual_recebimento)),
                        parte=nz.texto(_col(linha, cfg.manual_parte)),
                        civ=nz.texto(_col(linha, cfg.manual_civ)),
                        chamado=nz.texto(_col(linha, cfg.manual_chamado)),
                        observacao=nz.texto(_col(linha, cfg.manual_observacao)),
                        ano_origem=nz.texto(_col(linha, cfg.manual_ano)),
                        pendente=sch.status_e_pendente(status),
                        consolidado=cfg.manuais_consolidados,
                    )
                )
    finally:
        wb.close()

    historicos = [m for m in indice.manuais if not m.consolidado and m.pendente]
    if historicos:
        indice.avisos.append(
            Aviso(
                severidade="info",
                arquivo=sch.ARQ_REEMBOLSOS,
                aba="GUIAS 2025",
                mensagem=(
                    f"{len(historicos)} pendências manuais da aba de 2025 "
                    f"({nz.moeda(sum(m.valor for m in historicos))}) "
                    f"reaparecem na aba de 2026, que é a lista viva. Contei "
                    f"apenas a de 2026, para não duplicar."
                ),
            )
        )


def _avisar_notas_sem_baixa(indice: Indice) -> None:
    """Avisa sobre notas emitidas cuja coluna de recebimento ficou vazia.

    Nao e a mesma coisa que uma nota marcada PENDENTE, e o sistema nunca soma
    as duas caladamente. Ver pergunta 12 do mapeamento.
    """
    sem_baixa = [
        n
        for n in indice.notas
        if not n.bloco_secundario and n.estado == models_ESTADO_SEM_BAIXA
    ]
    if not sem_baixa:
        return
    abas = sorted({n.aba.strip() for n in sem_baixa})
    indice.avisos.append(
        Aviso(
            severidade="atencao",
            arquivo="Faturamento (ambos os CNPJs)",
            aba=", ".join(abas),
            mensagem=(
                f"{len(sem_baixa)} notas emitidas estão com a coluna de "
                f"recebimento vazia, somando "
                f"{nz.moeda(sum(n.valor for n in sem_baixa))}. Não estão "
                f"marcadas PENDENTE, porque o marcador só passou a ser usado "
                f"em julho de 2026. Podem ser cobranças vivas ou baixas que "
                f"ninguém registrou. Pergunta 12 do mapeamento, em aberto."
            ),
        )
    )


# ---------------------------------------------------------------------------
# Carga completa
# ---------------------------------------------------------------------------


def carregar(base: str) -> Indice:
    """Le os cinco arquivos e devolve o indice em memoria."""
    indice = Indice()

    for entidade in sch.ENTIDADES.values():
        _ler_faturamento(base, entidade, indice)
    _ler_razoes(base, indice)
    _ler_reembolsos(base, indice)
    _avisar_notas_sem_baixa(indice)

    for arquivo in sch.ARQUIVOS_ESPERADOS:
        caminho = _caminho(base, arquivo)
        if os.path.exists(caminho):
            stat = os.stat(caminho)
            indice.arquivos_lidos[arquivo] = dt.datetime.fromtimestamp(
                stat.st_mtime
            ).strftime("%d/%m/%Y %H:%M")
        else:
            indice.avisos.append(
                Aviso("critico", arquivo, None, "Arquivo nao encontrado na pasta.")
            )

    indice.carregado_em = dt.datetime.now()
    return indice


# ---------------------------------------------------------------------------
# Cache com invalidacao por mtime
# ---------------------------------------------------------------------------


class Repositorio:
    """Mantem o indice em memoria e relê quando algum arquivo muda."""

    def __init__(self, base: str) -> None:
        self.base = base
        self._indice: Optional[Indice] = None
        self._assinatura: Optional[str] = None
        self._lock = threading.Lock()

    def _assinatura_atual(self) -> str:
        partes = []
        for arquivo in sch.ARQUIVOS_ESPERADOS:
            caminho = _caminho(self.base, arquivo)
            if os.path.exists(caminho):
                stat = os.stat(caminho)
                partes.append(f"{arquivo}:{stat.st_mtime_ns}:{stat.st_size}")
            else:
                partes.append(f"{arquivo}:ausente")
        return hashlib.sha256("|".join(partes).encode()).hexdigest()

    def indice(self, forcar: bool = False) -> Indice:
        with self._lock:
            atual = self._assinatura_atual()
            if forcar or self._indice is None or atual != self._assinatura:
                self._indice = carregar(self.base)
                self._assinatura = atual
            return self._indice

    def recarregar(self) -> Indice:
        return self.indice(forcar=True)
