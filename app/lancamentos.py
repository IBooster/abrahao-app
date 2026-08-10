# -*- coding: utf-8 -*-
"""Motor de lancamentos: escreve nas planilhas do escritorio.

Segundo modulo autorizado a escrever, e o unico que altera CONTEUDO de
planilha. O outro (arquivos.py) so troca o arquivo inteiro.

Como funciona, e a ordem importa:

    1. propor()  monta uma Proposta dizendo exatamente arquivo, aba, linha e
                 celula que serao tocados, e o que falta perguntar. NAO escreve.
    2. a usuaria confere e confirma.
    3. aplicar() guarda copia do arquivo, escreve so as celulas da proposta e
                 registra na auditoria.

Regras que o codigo garante, nao a boa intencao:

    - so escreve em celula VAZIA, ou na coluna de recebimento de uma nota que
      a propria proposta identificou. Nunca sobrescreve valor por engano.
    - nunca escreve na aba 2026 (o financeiro preenche a mao no fechamento)
    - nunca escreve no arquivo da DRE (o openpyxl perde os desenhos dele)
    - nunca escreve em linha de bloco secundario
    - se qualquer conferencia falhar, nada e gravado

O que o openpyxl faz com o arquivo, medido nestes arquivos: preserva formula,
formato de moeda, largura de coluna, negrito e paineis congelados; perde o
valor em cache das formulas, que o Excel recalcula ao abrir.
"""

from __future__ import annotations

import datetime as dt
import os
import uuid
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl

from . import arquivos as arq
from .domain import normalize as nz
from .domain import schema as sch
from .domain.models import ESTADO_PENDENTE, ESTADO_SEM_BAIXA, Indice, Nota


class LancamentoRecusado(Exception):
    """A operacao nao passou na conferencia. Nada foi gravado."""


# Arquivos onde o sistema nunca escreve, e por que.
PROIBIDOS = {
    sch.ARQ_DRE_BMG: (
        "a DRE e um modelo mensal montado a mao, e salvar por aqui apagaria "
        "os 8 desenhos que ela tem"
    ),
}

# Abas onde o sistema nunca escreve, mesmo em arquivo permitido.
ABAS_PROIBIDAS = {
    sch.MATRIZ.aba: (
        "a matriz mensal e preenchida a mao pelo financeiro na conciliacao e "
        "no fechamento de cada mes"
    ),
}


# ---------------------------------------------------------------------------
# Proposta
# ---------------------------------------------------------------------------


@dataclass
class Celula:
    ref: str  # "B24"
    coluna: str  # rotulo humano: "Cliente"
    valor: Any
    exibicao: str


@dataclass
class Alvo:
    arquivo: str
    aba: str
    linha: int
    acao: str  # "nova linha" | "preenche coluna"
    celulas: list[Celula] = field(default_factory=list)


@dataclass
class Proposta:
    token: str
    tipo: str
    resumo: str
    alvos: list[Alvo] = field(default_factory=list)
    faltando: list[str] = field(default_factory=list)
    inferido: dict[str, str] = field(default_factory=dict)
    avisos: list[str] = field(default_factory=list)
    dados: dict[str, Any] = field(default_factory=dict)

    @property
    def pronta(self) -> bool:
        return not self.faltando and bool(self.alvos)


def _token() -> str:
    return uuid.uuid4().hex[:12]


# ---------------------------------------------------------------------------
# Apoio
# ---------------------------------------------------------------------------


def _entidade_do_cliente(ix: Indice, cliente: str) -> tuple[Optional[str], Optional[str]]:
    """Descobre por qual CNPJ o cliente costuma ser faturado.

    Primeiro pelo historico real de notas; depois pelo mapa da aba NOTAS.
    Devolve (codigo, explicacao) para a usuaria poder discordar.
    """
    notas = [
        n for n in ix.notas
        if n.emitida and not n.bloco_secundario and nz.cliente_bate(cliente, n.cliente)
    ]
    if notas:
        recente = max(notas, key=lambda n: (n.ano, n.mes, n.linha))
        ent = sch.ENTIDADES[recente.entidade]
        return recente.entidade, (
            f"pelo histórico: a última nota de {recente.cliente} "
            f"({recente.numero}) saiu por {ent.nome}"
        )

    destino = ix.mapa_cnpj.get(nz.radical_cliente(cliente))
    if destino == "rafaela":
        return "rafaela", "pela aba NOTAS, esse cliente migrou para o CNPJ da Rafaela"
    if destino == "principal":
        return "principal", "pela aba NOTAS, esse cliente permanece no CNPJ principal"
    return None, None


def _proporcao_liquido(
    ix: Indice, cliente: str, entidade: str
) -> tuple[Optional[float], int, dict]:
    """Retencao praticada com esse cliente, se as notas recentes concordarem.

    Arredonda a 4 casas de proposito: as planilhas trazem ruido na sexta casa
    (0,938498 e 0,938502 sao a mesma aliquota de 0,9385, arredondada em reais
    diferentes).

    Exige concordancia das mais recentes, nao a media. A FLAPA, por exemplo,
    teve cinco retencoes diferentes ao longo de 2026 - nesse caso devolve None
    e o sistema pergunta, em vez de chutar.
    """
    notas = [
        n for n in ix.notas
        if n.entidade == entidade and n.emitida and not n.bloco_secundario
        and nz.cliente_bate(cliente, n.cliente)
        and n.valor_bruto and n.valor_liquido and n.valor_bruto > 0
    ]
    if not notas:
        return None, 0, {}
    notas.sort(key=lambda n: (n.ano, n.mes, n.linha))
    recentes = notas[-6:]
    contagem: dict[float, int] = {}
    for n in recentes:
        r = round(n.valor_liquido / n.valor_bruto, 4)
        contagem[r] = contagem.get(r, 0) + 1

    dominante, vezes = max(contagem.items(), key=lambda x: x[1])
    # Precisa ser claramente a regra da casa, nao so a mais frequente por um
    # voto. O BMG usa 0,9385 nos honorarios e 0,9481 no Projeto Visitas; a
    # FLAPA trocou de retencao cinco vezes em 2026 e nao tem regra.
    if vezes * 3 < len(recentes) * 2:
        return None, len(recentes), contagem
    return dominante, len(recentes), contagem


def _proporcao_da_casa(ix: Indice, entidade: str) -> tuple[Optional[float], float]:
    """Retencao que o CNPJ pratica na maioria das notas.

    Serve para cliente novo, que nao tem historico proprio. Na sociedade
    principal sao 6,15% em 93% das notas; na Rafaela, que e do Simples, a
    maioria sai sem retencao. E chute com lastro, e a proposta diz isso.
    """
    notas = [
        n for n in ix.notas
        if n.entidade == entidade and n.emitida and not n.bloco_secundario
        and n.valor_bruto and n.valor_liquido and n.valor_bruto > 0
    ]
    if len(notas) < 10:
        return None, 0.0
    contagem: dict[float, int] = {}
    for n in notas:
        r = round(n.valor_liquido / n.valor_bruto, 4)
        contagem[r] = contagem.get(r, 0) + 1
    dominante, vezes = max(contagem.items(), key=lambda x: x[1])
    fatia = vezes / len(notas)
    if fatia < 0.6:
        return None, fatia
    return dominante, fatia


def _aba_do_mes(ix: Indice, entidade: str, mes: int, ano: int) -> Optional[str]:
    for n in ix.notas:
        if n.entidade == entidade and n.mes == mes and n.ano == ano:
            return n.aba
    return None


def _fim_do_bloco_principal(ix: Indice, entidade: str, aba: str) -> int:
    """Primeira linha livre depois do bloco principal daquela aba."""
    principais = [
        n.linha for n in ix.notas
        if n.entidade == entidade and n.aba == aba and not n.bloco_secundario
    ]
    return (max(principais) + 1) if principais else sch.COLS_FAT.primeira_linha


def _tem_bloco_secundario_em(ix: Indice, entidade: str, aba: str, linha: int) -> bool:
    return any(
        n.entidade == entidade and n.aba == aba and n.bloco_secundario and n.linha <= linha
        for n in ix.notas
    )


def _letra(coluna: int) -> str:
    return openpyxl.utils.get_column_letter(coluna)


# ---------------------------------------------------------------------------
# Operacao: nota emitida
# ---------------------------------------------------------------------------


def propor_nota_emitida(
    ix: Indice,
    cliente: Optional[str] = None,
    valor_bruto: Optional[float] = None,
    valor_liquido: Optional[float] = None,
    numero: Optional[str] = None,
    entidade: Optional[str] = None,
    competencia: Optional[str] = None,
    observacoes: Optional[str] = None,
    data_emissao: Optional[dt.date] = None,
    **_: Any,
) -> Proposta:
    """Registra uma nota emitida, ainda nao recebida.

    Escreve UMA linha na aba do mes, no arquivo do CNPJ correto, com a coluna
    de recebimento marcada PENDENTE. Nao mexe em caixa: dinheiro nao entrou.
    """
    p = Proposta(token=_token(), tipo="nota_emitida", resumo="")

    if not cliente:
        p.faltando.append("Para qual cliente?")
    if valor_bruto is None and valor_liquido is None:
        p.faltando.append("Qual o valor da nota?")
    if p.faltando:
        return p

    codigo = None
    if entidade:
        k = nz.chave(entidade)
        for c, e in sch.ENTIDADES.items():
            if k == c or any(a in k for a in e.apelidos):
                codigo = c
                break
    if not codigo:
        codigo, porque = _entidade_do_cliente(ix, cliente)
        if codigo and porque:
            p.inferido["CNPJ"] = f"{sch.ENTIDADES[codigo].nome} — {porque}"
    if not codigo:
        p.faltando.append(
            "Esse faturamento sai pela sociedade principal ou pela individual "
            "(Rafaela)? Não achei histórico desse cliente."
        )
        return p

    if valor_liquido is None:
        prop, quantas, contagem = _proporcao_liquido(ix, cliente, codigo)

        # Cliente novo nao tem historico. Em vez de deixar a usuaria no vacuo,
        # oferece a retencao que o CNPJ pratica na maioria das notas, dizendo
        # que e padrao da casa e nao do cliente.
        if prop is None and not quantas:
            casa, fatia = _proporcao_da_casa(ix, codigo)
            if casa is not None:
                valor_liquido = round(valor_bruto * casa, 2)
                p.inferido["Valor líquido"] = (
                    f"{nz.moeda(valor_liquido)} — {cliente} não tem nota "
                    f"anterior, então usei a retenção padrão de "
                    f"{sch.ENTIDADES[codigo].nome}: {(1 - casa) * 100:.2f}%, "
                    f"praticada em {fatia * 100:.0f}% das notas desse CNPJ. "
                    f"Confira antes de confirmar."
                )
                p.avisos.append(
                    "O valor líquido saiu do padrão do CNPJ, não do contrato "
                    "deste cliente. Se a retenção dele for outra, me diga o "
                    "líquido correto."
                )
                prop = casa

        if prop is None:
            if quantas:
                taxas = ", ".join(
                    f"{(1 - r) * 100:.2f}%" for r in sorted(contagem, reverse=True)
                )
                motivo = (
                    f"as últimas {quantas} notas desse cliente usaram retenções "
                    f"diferentes ({taxas}) e não dá para eleger uma"
                )
            else:
                motivo = "não tenho nota anterior desse cliente para me basear"
            p.faltando.append(
                f"Qual o valor líquido? O bruto é {nz.moeda(valor_bruto)}, mas "
                f"{motivo}."
            )
            return p

        # Quando o padrao da casa ja foi aplicado acima, a explicacao correta
        # ja esta em p.inferido - nao sobrescrever com a do historico do
        # cliente, que nesse caso nem existe.
        if "Valor líquido" not in p.inferido:
            valor_liquido = round(valor_bruto * prop, 2)
            retido = valor_bruto - valor_liquido
            nota_extra = ""
            if len(contagem) > 1:
                outras = ", ".join(
                    f"{(1 - r) * 100:.2f}%" for r in sorted(contagem) if r != prop
                )
                nota_extra = (
                    f"; atenção, "
                    f"{sum(v for r, v in contagem.items() if r != prop)} "
                    f"das últimas {quantas} usaram {outras}"
                )
            p.inferido["Valor líquido"] = (
                f"{nz.moeda(valor_liquido)} — retenção de "
                f"{(1 - prop) * 100:.2f}% ({nz.moeda(retido)}), praticada em "
                f"{contagem[prop]} das últimas {quantas} notas{nota_extra}"
            )
    if valor_bruto is None:
        valor_bruto = valor_liquido

    hoje = data_emissao or dt.date.today()
    aba = _aba_do_mes(ix, codigo, hoje.month, hoje.year)
    if aba is None:
        p.faltando.append(
            f"Não existe aba de {nz.rotulo_mes(hoje.month, hoje.year)} no "
            f"arquivo de {sch.ENTIDADES[codigo].nome}. Crie a aba no Excel "
            f"primeiro: o sistema não cria aba."
        )
        return p

    linha = _fim_do_bloco_principal(ix, codigo, aba)
    if _tem_bloco_secundario_em(ix, codigo, aba, linha):
        p.faltando.append(
            f"A linha {linha} da aba {aba.strip()} esbarra num bloco de layout "
            f"diferente. Não vou escrever às cegas ali."
        )
        return p

    if not numero:
        p.avisos.append(
            "Sem o número da NF, gravo a linha com o número em branco e ela "
            "fica como prevista. Assim que o Omie emitir, me diga o número "
            "que eu preencho."
        )

    if not competencia:
        competencia = dt.date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        p.inferido["Competência"] = f"{competencia} — primeiro dia do mês corrente"

    c = sch.COLS_FAT
    arquivo = sch.ENTIDADES[codigo].arquivo
    celulas = [
        Celula(f"{_letra(c.nf)}{linha}", "NF", numero, numero or "(em branco)"),
        Celula(f"{_letra(c.cliente)}{linha}", "Cliente", cliente, cliente),
        Celula(f"{_letra(c.valor_bruto)}{linha}", "Valor bruto", valor_bruto, nz.moeda(valor_bruto)),
        Celula(f"{_letra(c.valor_liquido)}{linha}", "Valor líquido", valor_liquido, nz.moeda(valor_liquido)),
        Celula(f"{_letra(c.referencia)}{linha}", "Referência", competencia, competencia),
        Celula(f"{_letra(c.observacoes)}{linha}", "Observações", observacoes, observacoes or "-"),
    ]
    if numero:
        celulas.append(
            Celula(f"{_letra(c.data_recebimento)}{linha}", "Data de recebimento",
                   sch.MARCADOR_PENDENTE, "PENDENTE")
        )

    p.alvos = [Alvo(arquivo=arquivo, aba=aba, linha=linha, acao="nova linha", celulas=celulas)]
    p.dados = {
        "cliente": cliente, "entidade": codigo, "aba": aba, "linha": linha,
        "valor_liquido": valor_liquido,
    }
    p.resumo = (
        f"Registrar nota de {nz.moeda(valor_liquido)} líquidos para {cliente}, "
        f"por {sch.ENTIDADES[codigo].nome}, na aba {aba.strip()}. "
        f"{'Fica marcada PENDENTE: ' if numero else 'Fica como prevista: '}"
        f"o dinheiro ainda não entrou, então nada é lançado no caixa."
    )
    return p


# ---------------------------------------------------------------------------
# Operacao: recebimento
# ---------------------------------------------------------------------------


def propor_recebimento(
    ix: Indice,
    cliente: Optional[str] = None,
    numero: Optional[str] = None,
    valor: Optional[float] = None,
    conta: Optional[str] = None,
    data: Optional[dt.date] = None,
    **_: Any,
) -> Proposta:
    """Da baixa numa nota e lanca a entrada no razao do banco.

    Duas escritas em arquivos diferentes, e as duas precisam acontecer: sem a
    linha no razao o dinheiro nao aparece no caixa; sem a data na coluna G a
    nota continua sendo cobrada.
    """
    p = Proposta(token=_token(), tipo="recebimento", resumo="")

    if not cliente and not numero:
        p.faltando.append("De qual cliente, ou qual o número da nota?")
        return p

    abertas = [
        n for n in ix.notas
        if not n.bloco_secundario and n.emitida
        and n.estado in (ESTADO_PENDENTE, ESTADO_SEM_BAIXA)
        and (numero is None or (n.numero or "").strip() == numero.strip())
        and (cliente is None or nz.cliente_bate(cliente, n.cliente))
    ]
    if not abertas:
        p.faltando.append(
            f"Não achei nota em aberto para {numero or cliente}. "
            f"Confira o número, ou pergunte 'o que {cliente or 'esse cliente'} "
            f"ainda deve'."
        )
        return p

    if valor is not None:
        candidatas = [n for n in abertas if abs(n.valor - valor) < 0.01]
        if candidatas:
            abertas = candidatas

    if len(abertas) > 1:
        lista = "; ".join(
            f"NF {n.numero} de {nz.moeda(n.valor)} ({nz.rotulo_mes(n.mes, n.ano)})"
            for n in abertas[:6]
        )
        p.faltando.append(f"Achei {len(abertas)} notas em aberto. Qual delas? {lista}")
        return p

    nota = abertas[0]
    quando = data or dt.date.today()

    c = _resolver_conta_recebimento(ix, nota, conta, p)
    if c is None:
        return p

    cols = sch.COLS_FAT
    ent = sch.ENTIDADES[nota.entidade]
    alvo_nota = Alvo(
        arquivo=ent.arquivo, aba=nota.aba, linha=nota.linha, acao="preenche coluna",
        celulas=[Celula(
            f"{_letra(cols.data_recebimento)}{nota.linha}", "Data de recebimento",
            quando, nz.data_br(quando),
        )],
    )

    linha_razao = _proxima_linha_razao(ix, c)
    alvo_razao = Alvo(
        arquivo=sch.ARQ_FLUXO, aba=c.aba, linha=linha_razao, acao="nova linha",
        celulas=[
            Celula(f"{_letra(c.remetente)}{linha_razao}", "Remetente", nota.cliente, nota.cliente or "-"),
            Celula(f"{_letra(c.contrato)}{linha_razao}", "Contrato", nota.cliente, nota.cliente or "-"),
            Celula(f"{_letra(c.descricao)}{linha_razao}", "Descrição", "FATURAMENTO", "FATURAMENTO"),
            Celula(f"{_letra(c.entrada)}{linha_razao}", "Entrada", nota.valor, nz.moeda(nota.valor)),
            Celula(f"{_letra(c.data)}{linha_razao}", c.nome_col_data.title(), quando, nz.data_br(quando)),
            Celula(f"{_letra(c.banco)}{linha_razao}", "Banco", c.marcadores[0], c.marcadores[0]),
            Celula(f"{_letra(c.historico)}{linha_razao}", "Histórico",
                   f"NF {nota.numero}" + (f" - {nota.observacoes}" if nota.observacoes else ""),
                   f"NF {nota.numero}"),
        ],
    )

    p.alvos = [alvo_nota, alvo_razao]
    p.dados = {"nf": nota.numero, "valor": nota.valor, "conta": c.codigo}
    p.resumo = (
        f"Dar baixa na NF {nota.numero} de {nota.cliente}, "
        f"{nz.moeda(nota.valor)}, recebida em {nz.data_br(quando)} no "
        f"{c.rotulo}. Duas escritas: a data na aba {nota.aba.strip()} e a "
        f"entrada no razão do {c.rotulo}."
    )
    p.avisos.append(
        "A matriz mensal do fluxo de caixa não é tocada: quem preenche é o "
        "financeiro, na conciliação do mês."
    )
    return p


def _resolver_conta_recebimento(ix: Indice, nota: Nota, conta: Optional[str], p: Proposta):
    if conta:
        k = nz.chave(conta)
        for x in sch.CONTAS.values():
            if k == x.codigo or any(nz.chave(a) == k for a in x.apelidos):
                return x
        p.faltando.append(f"Não conheço a conta '{conta}'.")
        return None

    entradas = [
        l for l in ix.lancamentos
        if not l.espelho and l.entrada > 0 and nz.cliente_bate(nota.cliente or "", l.contrato)
    ]
    if entradas:
        recente = max(entradas, key=lambda l: (l.data or dt.date.min))
        x = sch.CONTAS[recente.conta_efetiva]
        p.inferido["Conta"] = (
            f"{x.rotulo} — foi onde entrou o último recebimento desse cliente, "
            f"em {nz.data_br(recente.data)}"
        )
        return x

    p.faltando.append(
        "Em qual conta o dinheiro entrou? Não tenho histórico desse cliente "
        "para inferir."
    )
    return None


def _proxima_linha_razao(ix: Indice, c: sch.Conta) -> int:
    linhas = [l.linha for l in ix.lancamentos if l.aba == c.aba]
    return (max(linhas) + 1) if linhas else 2


# ---------------------------------------------------------------------------
# Aplicacao
# ---------------------------------------------------------------------------

OPERACOES = {
    "nota_emitida": propor_nota_emitida,
    "recebimento": propor_recebimento,
}


def propor(ix: Indice, tipo: str, dados: dict) -> Proposta:
    fn = OPERACOES.get(tipo)
    if fn is None:
        raise LancamentoRecusado(f"Não sei fazer '{tipo}'.")
    return fn(ix, **dados)


def _conferir_alvo(alvo: Alvo) -> None:
    if alvo.arquivo in PROIBIDOS:
        raise LancamentoRecusado(
            f"Não escrevo em {alvo.arquivo}: {PROIBIDOS[alvo.arquivo]}."
        )
    if alvo.aba.strip() in ABAS_PROIBIDAS:
        raise LancamentoRecusado(
            f"Não escrevo na aba {alvo.aba}: {ABAS_PROIBIDAS[alvo.aba.strip()]}."
        )
    if not alvo.celulas:
        raise LancamentoRecusado("Proposta sem célula nenhuma para gravar.")


def aplicar(base: str, proposta: Proposta, usuario: str) -> dict:
    """Grava a proposta. Guarda copia antes e registra na auditoria."""
    if not proposta.pronta:
        raise LancamentoRecusado(
            "Faltam dados: " + "; ".join(proposta.faltando)
        )
    for alvo in proposta.alvos:
        _conferir_alvo(alvo)

    por_arquivo: dict[str, list[Alvo]] = {}
    for alvo in proposta.alvos:
        por_arquivo.setdefault(alvo.arquivo, []).append(alvo)

    backups: dict[str, str] = {}
    gravadas: list[str] = []

    for arquivo, alvos in por_arquivo.items():
        caminho = os.path.join(base, arquivo)
        if not os.path.exists(caminho):
            raise LancamentoRecusado(f"Arquivo não está no servidor: {arquivo}")

        copia = arq._guardar_anterior(base, arquivo)
        if copia:
            backups[arquivo] = os.path.basename(copia)

        wb = openpyxl.load_workbook(caminho)
        try:
            for alvo in alvos:
                ws = None
                for nome in wb.sheetnames:
                    if nz.chave(nome) == nz.chave(alvo.aba):
                        ws = wb[nome]
                        break
                if ws is None:
                    raise LancamentoRecusado(f"Aba {alvo.aba} não existe em {arquivo}.")

                for cel in alvo.celulas:
                    atual = ws[cel.ref].value
                    if atual is not None and alvo.acao == "nova linha":
                        raise LancamentoRecusado(
                            f"A célula {cel.ref} da aba {alvo.aba} já tem "
                            f"conteúdo ({atual!r}). Não sobrescrevo: a planilha "
                            f"mudou desde que montei a proposta."
                        )
                    if cel.valor is not None:
                        ws[cel.ref] = cel.valor
                        gravadas.append(f"{alvo.aba}!{cel.ref}")

            temporario = caminho + ".parcial"
            wb.save(temporario)
        finally:
            wb.close()
        os.replace(temporario, caminho)

    arq._registrar(base, {
        "quando": dt.datetime.now().isoformat(timespec="seconds"),
        "usuario": usuario,
        "operacao": proposta.tipo,
        "token": proposta.token,
        "resumo": proposta.resumo,
        "celulas": gravadas,
        "arquivos": list(por_arquivo),
        "backups": backups,
        "dados": proposta.dados,
    })

    return {"celulas": gravadas, "backups": backups, "arquivos": list(por_arquivo)}
