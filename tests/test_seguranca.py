# -*- coding: utf-8 -*-
"""Testes de seguranca da Fase 2.

O compromisso e que o sistema NUNCA altera o conteudo de uma planilha: nao
abre celula, nao muda formula, nao acrescenta linha. Isso nao pode depender de
disciplina de quem programa; precisa falhar o teste se alguem introduzir um
caminho de escrita.

Existe UMA excecao, e ela e estreita de proposito: app/arquivos.py substitui
um arquivo inteiro por outro que a usuaria enviou. Isso existe porque no
Railway nao ha OneDrive e alguem precisa colocar as planilhas la. Substituir o
arquivo todo nao e alterar o dado contabil - o conteudo nunca passa pelo
sistema.

Os testes abaixo cercam essa excecao: nenhum outro modulo escreve, e esse so
escreve nos cinco nomes conhecidos, dentro da pasta configurada.

Rodar:  python -m pytest tests -v
"""

from __future__ import annotations

import ast
import os
import sys
from pathlib import Path

import pytest

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

APP = RAIZ / "app"

# Metodos que gravam arquivo independentemente de quem os chame.
# 'save' cobre Workbook.save, que e o risco real neste projeto.
METODOS_SEMPRE_PROIBIDOS = {"save", "truncate", "writelines"}

# Metodos de filesystem que so contam quando chamados sobre os modulos abaixo.
# Sem esse recorte, str.replace e list.remove viram falso positivo.
METODOS_DE_FILESYSTEM = {
    "remove",
    "unlink",
    "rmtree",
    "rename",
    "replace",
    "mkdir",
    "makedirs",
    "copy",
    "copy2",
    "copyfile",
    "move",
}
MODULOS_DE_FILESYSTEM = {"os", "shutil", "pathlib", "Path"}

MODOS_DE_ESCRITA = ("w", "a", "x", "+")


# Os unicos modulos autorizados a escrever em disco:
#   arquivos.py    troca o arquivo inteiro (envio das planilhas)
#   lancamentos.py altera celula, sempre depois de proposta confirmada
MODULOS_DE_ESCRITA = {APP / "arquivos.py", APP / "lancamentos.py"}


def _arquivos_python(incluir_guarda: bool = False) -> list[Path]:
    todos = [p for p in APP.rglob("*.py") if "__pycache__" not in str(p)]
    if incluir_guarda:
        return todos
    return [p for p in todos if p not in MODULOS_DE_ESCRITA]


def _raiz_do_receptor(no: ast.AST) -> str | None:
    """Nome da base de uma cadeia de atributos: os.path.join -> 'os'."""
    atual = no
    while isinstance(atual, ast.Attribute):
        atual = atual.value
    if isinstance(atual, ast.Name):
        return atual.id
    return None


def test_nenhuma_chamada_de_escrita_fora_do_modulo_de_guarda():
    """So app/arquivos.py pode gravar ou apagar arquivo."""
    achados = []
    for caminho in _arquivos_python():
        arvore = ast.parse(caminho.read_text(encoding="utf-8"), filename=str(caminho))
        for no in ast.walk(arvore):
            if not isinstance(no, ast.Call) or not isinstance(no.func, ast.Attribute):
                continue
            metodo = no.func.attr
            local = f"{caminho.relative_to(RAIZ)}:{no.lineno}"

            if metodo in METODOS_SEMPRE_PROIBIDOS:
                achados.append(f"{local} -> {metodo}()")
                continue

            if metodo in METODOS_DE_FILESYSTEM:
                raiz = _raiz_do_receptor(no.func.value)
                if raiz in MODULOS_DE_FILESYSTEM:
                    achados.append(f"{local} -> {raiz}.{metodo}()")

    assert not achados, (
        "Chamada de escrita fora de app/arquivos.py. Se precisar mesmo "
        "escrever, faca pelo modulo de guarda:\n  " + "\n  ".join(achados)
    )


def test_modulo_de_guarda_so_escreve_nos_arquivos_conhecidos():
    """O modulo de escrita nunca aceita nome de arquivo vindo de fora.

    O destino sai sempre do catalogo em schema.ARQUIVOS, depois de passar por
    os.path.basename. Sem isso, um nome como '../../etc/senha' gravaria fora
    da pasta.
    """
    from app import arquivos

    fonte = (APP / "arquivos.py").read_text(encoding="utf-8")
    assert "os.path.basename" in fonte, (
        "O modulo de escrita precisa reduzir o nome recebido a basename, "
        "senao aceita caminho e grava fora da pasta."
    )

    for tentativa in (
        "../fora.xlsx",
        "../../etc/passwd",
        "C:\\Windows\\System32\\qualquer.xlsx",
        "aleatorio.xlsx",
        "",
    ):
        with pytest.raises(arquivos.ArquivoRecusado):
            arquivos.conferir(tentativa, b"PK\x03\x04conteudo")


def test_modulo_de_guarda_recusa_o_que_nao_e_planilha():
    from app import arquivos
    from app.domain import schema

    nome = schema.ARQ_REEMBOLSOS
    with pytest.raises(arquivos.ArquivoRecusado):
        arquivos.conferir(nome, b"")
    with pytest.raises(arquivos.ArquivoRecusado):
        arquivos.conferir(nome, b"isto nao e um xlsx")
    with pytest.raises(arquivos.ArquivoRecusado):
        arquivos.conferir(nome, b"PK\x03\x04" + b"\x00" * 200)


def test_upload_confere_abas_antes_de_substituir(tmp_path):
    """Arquivo com o nome certo mas conteudo errado nao substitui o bom."""
    import openpyxl

    from app import arquivos
    from app.domain import schema

    nome = schema.ARQ_REEMBOLSOS
    bom = tmp_path / nome
    bom.write_bytes(b"planilha boa que ja estava la")

    # Um xlsx valido, porem sem a aba obrigatoria GUIAS 2026.
    wb = openpyxl.Workbook()
    wb.active.title = "outra coisa"
    impostor = tmp_path / "impostor.xlsx"
    wb.save(str(impostor))
    wb.close()

    with pytest.raises(arquivos.ArquivoRecusado) as erro:
        arquivos.receber(str(tmp_path), nome, impostor.read_bytes(), "teste")

    assert "GUIAS 2026" in str(erro.value)
    assert bom.read_bytes() == b"planilha boa que ja estava la", (
        "O arquivo que ja existia foi tocado mesmo com o envio recusado."
    )


def test_upload_guarda_copia_antes_de_substituir(tmp_path):
    import openpyxl

    from app import arquivos
    from app.domain import schema

    nome = schema.ARQ_REEMBOLSOS
    anterior = tmp_path / nome
    anterior.write_bytes(b"versao anterior")

    wb = openpyxl.Workbook()
    wb.active.title = "GUIAS 2026"
    novo = tmp_path / "novo.xlsx"
    wb.save(str(novo))
    wb.close()

    recebido = arquivos.receber(
        str(tmp_path), nome, novo.read_bytes(), "teste"
    )

    assert recebido.substituiu is True
    assert recebido.backup, "Nenhuma copia foi guardada antes de substituir."

    copias = list((tmp_path / schema.PASTA_BACKUPS).glob("*.xlsx"))
    assert len(copias) == 1
    assert copias[0].read_bytes() == b"versao anterior"

    registros = arquivos.historico(str(tmp_path))
    assert len(registros) == 1
    assert registros[0]["arquivo"] == nome
    assert registros[0]["usuario"] == "teste"


def test_nenhum_open_em_modo_de_escrita():
    """Nenhum open() do app usa modo de escrita."""
    achados = []
    for caminho in _arquivos_python():
        arvore = ast.parse(caminho.read_text(encoding="utf-8"), filename=str(caminho))
        for no in ast.walk(arvore):
            if not isinstance(no, ast.Call):
                continue
            alvo = no.func
            nome = alvo.id if isinstance(alvo, ast.Name) else getattr(alvo, "attr", None)
            if nome not in ("open", "write_text", "write_bytes"):
                continue
            if nome in ("write_text", "write_bytes"):
                achados.append(f"{caminho.relative_to(RAIZ)}:{no.lineno} -> {nome}()")
                continue
            if len(no.args) >= 2 and isinstance(no.args[1], ast.Constant):
                modo = str(no.args[1].value)
                if any(m in modo for m in MODOS_DE_ESCRITA):
                    achados.append(
                        f"{caminho.relative_to(RAIZ)}:{no.lineno} -> open(..., {modo!r})"
                    )
    assert not achados, "open() em modo de escrita:\n  " + "\n  ".join(achados)


def test_workbooks_abertos_somente_leitura():
    """Toda chamada a load_workbook passa read_only=True."""
    from app.domain import loader

    fonte = Path(loader.__file__).read_text(encoding="utf-8")
    arvore = ast.parse(fonte)
    chamadas = [
        no
        for no in ast.walk(arvore)
        if isinstance(no, ast.Call)
        and getattr(no.func, "attr", None) == "load_workbook"
    ]
    assert chamadas, "Esperava encontrar load_workbook no loader."
    for chamada in chamadas:
        argumentos = {kw.arg: kw.value for kw in chamada.keywords}
        assert "read_only" in argumentos, (
            f"load_workbook na linha {chamada.lineno} sem read_only."
        )
        valor = argumentos["read_only"]
        assert isinstance(valor, ast.Constant) and valor.value is True, (
            f"load_workbook na linha {chamada.lineno} com read_only diferente de True."
        )


def test_nenhuma_rota_de_escrita_em_planilha():
    """As rotas POST existentes nao alteram planilha."""
    from app import main

    rotas_post = {
        rota.path
        for rota in main.app.routes
        if "POST" in getattr(rota, "methods", set())
    }
    permitidas = {
        "/entrar",
        "/sair",
        "/api/perguntar",
        "/api/recarregar",
        "/api/planilhas",   # recebe arquivo inteiro, nunca altera conteudo
        "/api/confirmar",   # aplica proposta que a usuaria confirmou na tela
        "/api/cancelar",
    }
    inesperadas = rotas_post - permitidas
    assert not inesperadas, f"Rota POST nao prevista na Fase 2: {inesperadas}"


def test_schema_declara_modo_somente_leitura():
    from app.domain import schema

    assert schema.MODO_SOMENTE_LEITURA is True


def test_motor_de_consulta_nao_escreve():
    """Nenhuma operacao de escrita vazou para o catalogo de consultas."""
    from app.domain import schema
    from app.queries import engine

    for nome in schema.OPERACOES_ESCRITA_PLANEJADAS:
        assert nome not in engine.CATALOGO, (
            f"A operacao de escrita '{nome}' esta no catalogo de consultas. "
            f"Consulta nunca altera arquivo."
        )


def test_arquivos_nao_mudam_apos_uso():
    """Le, consulta tudo, e confere que nenhum arquivo foi tocado no disco."""
    from app.config import CONFIG
    from app.domain import schema
    from app.domain.loader import carregar
    from app.queries import engine

    base = CONFIG.pasta_planilhas
    if not os.path.isdir(base):
        pytest.skip("Pasta de planilhas nao disponivel neste ambiente.")

    antes = {}
    for arquivo in schema.ARQUIVOS_ESPERADOS:
        caminho = os.path.join(base, arquivo)
        if os.path.exists(caminho):
            info = os.stat(caminho)
            antes[arquivo] = (info.st_mtime_ns, info.st_size)

    indice = carregar(base)
    for nome, consulta in engine.CATALOGO.items():
        parametros = {}
        if "cliente" in consulta.parametros:
            parametros["cliente"] = "BMG"
        if "conta" in consulta.parametros:
            parametros["conta"] = "santander"
        if "mes" in consulta.parametros:
            parametros["mes"] = 7
        engine.executar(indice, nome, parametros)

    for arquivo, esperado in antes.items():
        caminho = os.path.join(base, arquivo)
        info = os.stat(caminho)
        assert (info.st_mtime_ns, info.st_size) == esperado, (
            f"O arquivo {arquivo} mudou depois de consultar. A Fase 2 nao "
            f"pode alterar planilha."
        )
